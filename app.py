"""
ID Card Details Collector
--------------------------
A small Flask website that collects student ID card details
and automatically appends every submission as a new row in an
Excel file (id_card_data.xlsx) using openpyxl.
"""

import os
import re
import random
import secrets
from functools import wraps
from datetime import datetime

from flask import Flask, render_template, request, redirect, url_for, flash, send_file, Response
from openpyxl import Workbook, load_workbook
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address

app = Flask(__name__)

# IMPORTANT: generate your own key once by running this in a Python console:
#   import secrets; print(secrets.token_hex(32))
# then paste the result below instead of this placeholder.
app.secret_key = "8e3ced2986af5a078e8ef56dd134d0cca77f55186e172b90d25b46a0b29861df"

# ---- Basic auth credentials for the /download route ----
ADMIN_USERNAME = "kishor"
ADMIN_PASSWORD = "Kishor@2008"  # change this to something only you know

# ---- Rate limiting setup ----
limiter = Limiter(get_remote_address, app=app, default_limits=["50 per hour"])

# Excel file where every submission is stored
EXCEL_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "id_card_data.xlsx")
SHEET_NAME = "ID Card Details"
HEADERS = [
    "Name",
    "Class",
    "Date of Birth",
    "Blood Group",
    "Father's Name",
    "Address",
    "Mobile Number",
    "Photo no",
    "Submitted At",
    ]


def init_excel():
    """Create the Excel file with headers if it doesn't already exist."""
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = SHEET_NAME
        ws.append(HEADERS)
        for col_idx, header in enumerate(HEADERS, start=1):
            ws.column_dimensions[chr(64 + col_idx)].width = max(15, len(header) + 4)
        wb.save(EXCEL_FILE)


def append_to_excel(row):
    """Append one row of form data to the Excel sheet."""
    init_excel()
    wb = load_workbook(EXCEL_FILE)
    ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active
    ws.append(row)
    wb.save(EXCEL_FILE)


# ---- Basic auth helper functions (protects /download) ----
def check_auth(username, password):
    return username == ADMIN_USERNAME and password == ADMIN_PASSWORD


def authenticate():
    return Response(
        "Access denied. Admin login required.",
        401,
        {"WWW-Authenticate": 'Basic realm="Login Required"'},
    )


def requires_auth(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        auth = request.authorization
        if not auth or not check_auth(auth.username, auth.password):
            return authenticate()
        return f(*args, **kwargs)
    return decorated


@app.route("/", methods=["GET", "POST"])
@limiter.limit("10 per hour")
def index():
    if request.method == "POST":
        name = request.form.get("name", "").strip()
        student_class = request.form.get("student_class", "").strip()
        dob = request.form.get("dob", "").strip()
        if dob:
            try:
                dob=datetime.strptime(dob, "%Y-%m-%d").strftime("%d-%m-%Y")
            except ValueError:
                pass
        blood_group = request.form.get("blood_group", "").strip()
        father_name = request.form.get("father_name", "").strip()
        address = request.form.get("address", "").strip()
        mobile = request.form.get("mobile", "").strip()
        photo_no = request.form.get("photo_no", "").strip()

        errors = []
        if not name:
            errors.append("Please enter the name.")
        elif len(name) > 100:
            errors.append("Name is too long.")
        if not student_class:
            errors.append("Please enter the class.")
        if not dob:
            errors.append("Please enter the date of birth.")
        if not blood_group:
            errors.append("Please select the blood group.")
        if not father_name:
            errors.append("Please enter the father's name.")
        elif len(father_name) > 100:
            errors.append("Father's name is too long.")
        if not address:
            errors.append("Please enter the address.")
        elif len(address) > 300:
            errors.append("Address is too long.")
        if not re.fullmatch(r"\d{10}", mobile):
            errors.append("Please enter only 10 numbers for the mobile number.")
        if not photo_no:
            errors.append("Please enter the photo no.")

        if errors:
            for e in errors:
                flash(e)
            return render_template(
                "index.html",
                form=request.form,
                form_no=random.randint(1000, 9999),
            )

        append_to_excel(
            [
                name,
                student_class,
                dob,
                blood_group,
                father_name,
                address,
                mobile,
                photo_no,
                datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            ]
        )

        return redirect(url_for("success"))

    return render_template("index.html", form={}, form_no=random.randint(1000, 9999))


@app.route("/success")
def success():
    return render_template("success.html")


@app.route("/download")
@requires_auth
def download():
    """Only accessible with the admin username/password (basic auth)."""
    init_excel()
    return send_file(EXCEL_FILE, as_attachment=True)


if __name__ == "__main__":
    init_excel()
    app.run(debug=False)